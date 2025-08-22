import os
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

def insert_images_to_fai_template(template_path, image_folder, sheet_name="PICTURE"):
    # --------------------------
    # 配置参数（按新列数调整）
    # --------------------------
    col_width = 8.43  # 列宽（字符）
    row_height = 12.75  # 行高（磅）
    rows_per_box = 16  # 单个选框占用行数
    # 第一列：B:H（7列，索引2-8）
    first_col1, last_col1 = 2, 8  
    # 第二列：J:O（6列，索引10-15）
    first_col2, last_col2 = 10, 15  
    
    # --------------------------
    # 1. 打开Excel模板
    # --------------------------
    try:
        wb = openpyxl.load_workbook(template_path)
        ws = wb[sheet_name]
    except Exception as e:
        print(f"打开模板失败：{e}")
        return
    
    # --------------------------
    # 2. 收集图片路径
    # --------------------------
    image_extensions = ('.jpg', '.jpeg', '.png', '.bmp')
    image_paths = [
        os.path.join(image_folder, f) 
        for f in os.listdir(image_folder) 
        if f.lower().endswith(image_extensions)
    ]
    pic_count = len(image_paths)
    if pic_count == 0:
        print("未找到图片文件！")
        return
    
    # --------------------------
    # 3. 检测现有预留合并单元格
    # --------------------------
    reserve_ranges = []  # 存储预留区域的（起始行, 结束行, 起始列, 结束列）
    last_end_row = 0     # 最后一个预留区域的结束行
    
    # 遍历所有合并单元格
    for merged_range in ws.merged_cells.ranges:
        # 解析合并区域（格式如：B1:H16 或 J1:O16）
        min_col, min_row, max_col, max_row = merged_range.bounds
        row_count = max_row - min_row + 1
        col_count = max_col - min_col + 1
        
        # 判断是否符合预留区域规格（匹配列范围和行高）
        if (row_count == rows_per_box and 
            ((min_col == first_col1 and max_col == last_col1) or  # 第一列B:H
             (min_col == first_col2 and max_col == last_col2))):  # 第二列J:O
            reserve_ranges.append((min_row, max_row, min_col, max_col))
            if max_row > last_end_row:
                last_end_row = max_row
    
    # 若没有预留区域，从第9行开始（默认起始位置）
    if not reserve_ranges:
        last_end_row = 8  # 从第9行开始（8+1）
    
    reserve_count = len(reserve_ranges)
    print(f"检测到预留合并单元格：{reserve_count} 个")
    
    # --------------------------
    # 4. 按需新增合并单元格
    # --------------------------
    need_new = max(0, pic_count - reserve_count)
    new_start_row = last_end_row + 1  # 新增区域的起始行
    new_ranges = []
    
    for i in range(need_new):
        # 交替使用两列（第一列→第二列→换行）
        if i % 2 == 0:
            start_col, end_col = first_col1, last_col1  # B:H（7列）
        else:
            start_col, end_col = first_col2, last_col2  # J:O（6列）
        
        # 计算新增区域的行范围
        start_row = new_start_row
        end_row = start_row + rows_per_box - 1
        
        # 合并单元格
        merge_range = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
        ws.merge_cells(merge_range)
        
        # 设置行高和列宽
        for row in range(start_row, end_row + 1):
            ws.row_dimensions[row].height = row_height
        for col in range(start_col, end_col + 1):
            ws.column_dimensions[get_column_letter(col)].width = col_width
        
        # 记录新增区域
        new_ranges.append((start_row, end_row, start_col, end_col))
        print(f"新增合并单元格：{merge_range}")
        
        # 第二列新增后换行
        if i % 2 == 1:
            new_start_row = end_row + 1
    
    # 合并预留区域和新增区域
    all_ranges = reserve_ranges + new_ranges
    
    # --------------------------
    # 5. 插入图片
    # --------------------------
    # 清除现有图片（避免重复）
    ws._images = []  # 清空工作表中的图片
    
    # 插入图片到所有区域
    for i in range(pic_count):
        try:
            # 获取目标区域
            start_row, end_row, start_col, end_col = all_ranges[i]
            # 计算区域尺寸（像素，用于图片缩放）
            cell_width = ws.column_dimensions[get_column_letter(start_col)].width * 7.5  # 1字符≈7.5像素
            cell_height = ws.row_dimensions[start_row].height * 1.33  # 1磅≈1.33像素
            total_width = cell_width * (end_col - start_col + 1)  # 总宽度=列宽×列数
            total_height = cell_height * (end_row - start_row + 1)  # 总高度=行高×行数
            
            # 插入图片并缩放
            img = Image(image_paths[i])
            img.width = total_width
            img.height = total_height
            # 图片左上角对齐单元格左上角
            ws.add_image(img, anchor=f"{get_column_letter(start_col)}{start_row}")
            print(f"插入图片 {i+1}/{pic_count}：{os.path.basename(image_paths[i])}")
        except Exception as e:
            print(f"插入图片失败（第{i+1}张）：{e}")
    
    # --------------------------
    # 6. 保存文件
    # --------------------------
    try:
        output_path = os.path.splitext(template_path)[0] + "_已插入图片.xlsx"
        wb.save(output_path)
        print(f"\n处理完成！文件已保存至：{output_path}")
        print(f"总图片数：{pic_count}，使用预留区域：{min(reserve_count, pic_count)}，新增区域：{need_new}")
    except Exception as e:
        print(f"保存文件失败：{e}")

# --------------------------
# 执行脚本（修改以下路径）
# --------------------------
if __name__ == "__main__":
    # 替换为你的模板路径和图片文件夹路径
    TEMPLATE_PATH = r"E:\工作区\FAI模板.xlsx"  # FAI模板文件路径
    IMAGE_FOLDER = r"E:\工作区\Hops图片维护设备\1-客户资料\2025-7-31-Hellen发\N13161-122(一对一)\N13161-122(一对一)"  # 图片文件夹路径
    
    insert_images_to_fai_template(TEMPLATE_PATH, IMAGE_FOLDER)
    